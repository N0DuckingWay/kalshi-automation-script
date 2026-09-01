"""
File: reporter.py
Author: Zachary Hoffman
Last edited by: Zachary Hoffman

Purpose:
    Handles all Excel output for the bot. In production mode, appends a run-
    separator row followed by per-trade rows to a persistent trade_log.xlsx file
    so the full trading history accumulates across runs. In dev/sandbox mode,
    writes a fresh timestamped simulation file containing two sheets: one for
    simulated trades and one for all candidate pairs discovered (tradeable or not).
    All Excel formatting — column widths, color-coded status rows, number formats,
    frozen header rows — is applied via openpyxl.

Dependencies:
    Imports display_title from scanner.py and TradeSpec from strategy.py. Imports
    PROJECT_ROOT from config.py. Exports the TradeResult dataclass (consumed by
    trader.py) and the two public write functions (consumed by main.py).

Notes:
    The TradeResult dataclass is defined here (not in trader.py) because reporter.py
    is the authoritative consumer of trade outcomes — trader.py only needs to
    construct and return these objects.

    append_to_prod_log() coordinates concurrent writers (e.g. the scheduler and a
    manual invocation racing) via a sidecar advisory lock (fcntl.flock) around the
    load -> append -> save sequence, and saves atomically (tmp file + os.replace)
    so a crash mid-save can never truncate the accumulated trade history. If the
    lock cannot be acquired within _LOCK_TIMEOUT_SECONDS, this run's rows are never
    silently dropped — they're written to a standalone timestamped fallback file
    instead of touching the shared log. See BS-18 in CLAUDE.md's bug-sweep history.
"""
import fcntl
import logging
import os
import time
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import PROJECT_ROOT
from .scanner import display_title
from .strategy import TradeSpec

PROD_LOG_PATH = PROJECT_ROOT / "trade_log.xlsx"

# Sidecar lock file coordinating concurrent writers to PROD_LOG_PATH (e.g. the
# weekly scheduler and a manual run racing) — see _acquire_lock().
_LOCK_PATH = PROD_LOG_PATH.with_name(PROD_LOG_PATH.name + ".lock")
# Bounded acquire loop: a non-blocking flock() attempt polled at this interval,
# up to this many total seconds, rather than a blocking flock() call — so a
# wedged lock holder (crashed mid-hold, or just slow) can never hang this run
# forever. Deliberately short relative to a full bot run.
_LOCK_TIMEOUT_SECONDS = 30
_LOCK_POLL_SECONDS = 0.5

# Column definitions shared by both sheets
_TRADE_COLUMNS = [
    ("Date",              14),
    ("Time",              10),
    ("Market A",          45),
    ("Ticker A",          18),
    ("Market B",          45),
    ("Ticker B",          18),
    ("A Deadline",        12),
    ("B Deadline",        12),
    ("pA (YES ask)",      13),
    ("pB (YES ask)",      13),
    ("nA (NO ask)",       13),
    ("x — NO on A",       12),
    ("y — YES on B",      13),
    ("Total Cost ($)",    14),
    ("Min Profit ($)",    14),
    ("Profit Ratio (%)",  16),
    ("Status",            12),
    ("Notes",             30),
]

_HEADER_FILL_PROD = PatternFill("solid", fgColor="1F4E79")   # dark blue for prod
_HEADER_FILL_DEV  = PatternFill("solid", fgColor="375623")   # dark green for dev
_HEADER_FONT      = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL   = PatternFill("solid", fgColor="D6E4F7")   # light blue
_SUBHEADER_FONT   = Font(bold=True, color="1F4E79", size=10)
_THIN_BORDER      = Border(
    bottom=Side(style="thin", color="BFBFBF"),
)


@dataclass
class TradeResult:
    """
    Outcome record for a single attempted or simulated trade.

    Attributes:
        spec (TradeSpec): The trade specification that was executed or simulated.
        status (str): Execution outcome — "executed" for a real submitted order,
            "simulated" for a dry-run or dev-mode run, "failed" for a submission error,
            "rolled_back" if leg A filled but leg B failed and leg A was unwound,
            "rollback_failed" if the leg A unwind itself did not fill (orphaned
            position requiring manual review), or "manual_review" if leg B's fill
            state could not be determined (position lookup failed) and no
            automated rollback was attempted to avoid reversing a possible real fill.
        error (Optional[str]): Error message if status is "failed", "rolled_back",
            "rollback_failed", or "manual_review", otherwise None.
    """
    spec: TradeSpec
    status: str            # "executed" | "failed" | "simulated" | "rolled_back" | "rollback_failed" | "manual_review"
    error: str | None = None



def _apply_header_row(ws, fill: PatternFill) -> None:
    """Write and style the column header row."""
    for col_idx, (header, width) in enumerate(_TRADE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _result_to_row(result: TradeResult, run_ts: datetime) -> list:
    """
    Serialize a TradeResult to a flat list matching the _TRADE_COLUMNS column order.

    Extracts all fields needed for one Excel data row, formatting prices as rounded
    floats and datetimes as "YYYY-MM-DD" strings.

    Args:
        result (TradeResult): The trade result to serialize.
        run_ts (datetime): Timestamp of the current bot run, used to populate the
            Date and Time columns.

    Returns:
        list: Ordered list of 18 values, one per column in _TRADE_COLUMNS.
    """
    spec = result.spec
    pair = spec.pair
    mA   = display_title(pair.market_a)
    mB   = display_title(pair.market_b)

    def fmt_dt(dt) -> str:
        return dt.strftime("%Y-%m-%d") if dt else ""

    return [
        run_ts.strftime("%Y-%m-%d"),
        run_ts.strftime("%H:%M:%S"),
        mA,
        pair.market_a.ticker,
        mB,
        pair.market_b.ticker,
        fmt_dt(pair.market_a.close_time),
        fmt_dt(pair.market_b.close_time),
        round(pair.pA, 4),
        round(pair.pB, 4),
        round(pair.nA, 4),
        spec.x,
        spec.y,
        round(spec.total_cost, 2),
        round(spec.min_payoff, 2),
        round(spec.profit_ratio, 4),
        result.status,
        result.error or "",
    ]


def _apply_data_row_styles(ws, row_idx: int, status: str) -> None:
    """
    Apply background fill, border, and alignment styling to a single data row.

    Color-codes rows by trade status: green for "executed", blue for "simulated",
    red/orange for "failed", yellow for "rolled_back", strong red for
    "rollback_failed" and "manual_review", white for any unknown status.

    Args:
        ws: An openpyxl Worksheet object to apply styles to.
        row_idx (int): 1-based row index of the data row to style.
        status (str): Trade status string — "executed", "simulated", "failed",
            "rolled_back", "rollback_failed", or "manual_review".
    """
    status_colors = {
        "executed":        "E2EFDA",   # light green
        "simulated":       "EBF3FB",   # light blue
        "failed":          "FCE4D6",   # light red/orange
        "rolled_back":     "FFF2CC",   # light yellow — leg A unwound, no net position
        "rollback_failed": "F4B7B4",   # strong red — orphaned position, manual review
        "manual_review":   "F4B7B4",   # strong red — fill state unknown, manual review
    }
    fill_color = status_colors.get(status, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)
    for col in range(1, len(_TRADE_COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill   = fill
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def _apply_number_formats(ws, row_idx: int) -> None:
    """Apply currency/percentage formats to numeric columns."""
    # Columns: pA=9, pB=10, nA=11, TotalCost=14, MinProfit=15, ProfitRatio=16
    for col in (9, 10, 11):
        ws.cell(row=row_idx, column=col).number_format = "0.00%"
    for col in (14, 15):
        ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'
    ws.cell(row=row_idx, column=16).number_format = "0.00%"


def _write_separator_row(
    ws, run_ts: datetime, balance_before: float, balance_after: float, n_results: int
) -> None:
    """
    Append a styled run-separator row summarizing this run's balance change.

    Shared by the normal prod-log append path and the lock-timeout fallback path
    so both files carry the same run-summary banner.

    Args:
        ws: The openpyxl Worksheet to append to.
        run_ts (datetime): Timestamp of this run.
        balance_before (float): Account balance in dollars before this run's trades.
        balance_after (float): Account balance in dollars after this run's trades.
        n_results (int): Number of trade results in this run, shown in the banner.
    """
    sep_row = ws.max_row + 1
    sep_cell = ws.cell(row=sep_row, column=1,
                       value=f"── Run: {run_ts.strftime('%Y-%m-%d %H:%M')}  |  "
                             f"Balance before: ${balance_before:.2f}  →  "
                             f"after: ${balance_after:.2f}  |  "
                             f"{n_results} trade(s)")
    sep_cell.font = Font(italic=True, color="595959", size=9)
    sep_cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.merge_cells(
        start_row=sep_row, start_column=1,
        end_row=sep_row, end_column=len(_TRADE_COLUMNS)
    )


def _write_trade_rows(ws, results: list, run_ts: datetime) -> None:
    """
    Append one styled, color-coded data row per TradeResult.

    Shared by the normal prod-log append path and the lock-timeout fallback path.

    Args:
        ws: The openpyxl Worksheet to append to.
        results (list): List of TradeResult objects to write, in order.
        run_ts (datetime): Timestamp of this run, used for the Date/Time columns.
    """
    for result in results:
        row_idx = ws.max_row + 1
        row_data = _result_to_row(result, run_ts)
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _apply_data_row_styles(ws, row_idx, result.status)
        _apply_number_formats(ws, row_idx)


def _acquire_lock(lock_path: Path):
    """
    Acquire an exclusive advisory lock on lock_path, polling up to _LOCK_TIMEOUT_SECONDS.

    Uses a bounded loop of non-blocking fcntl.flock() attempts rather than a single
    blocking flock() call, so a lock held by a hung (but still alive) process can
    never wedge this call forever — the caller falls back to a standalone file
    instead of waiting indefinitely.

    Args:
        lock_path (Path): Path to the sidecar lock file. Created if it doesn't exist.

    Returns:
        The open file object holding the lock (caller must close it — via
        _release_lock — when done), or None if the timeout elapsed first.
    """
    lock_path.touch(exist_ok=True)
    fh = open(lock_path, "r+")
    deadline = time.monotonic() + _LOCK_TIMEOUT_SECONDS
    while True:
        try:
            fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            return fh
        except OSError:
            if time.monotonic() >= deadline:
                fh.close()
                return None
            time.sleep(_LOCK_POLL_SECONDS)


def _release_lock(lock_fh) -> None:
    """
    Release a lock acquired by _acquire_lock() and close its file handle.

    Deliberately does NOT unlink the lock file: removing a flock'd file while
    another process might be about to open/lock that same path is a TOCTOU race
    (the other process could end up locking a different, newly-created inode of
    the same name, defeating the lock entirely). The sidecar file is small and
    harmless to keep around indefinitely — it's gitignored.

    Args:
        lock_fh: The open file object returned by _acquire_lock().
    """
    try:
        fcntl.flock(lock_fh.fileno(), fcntl.LOCK_UN)
    finally:
        lock_fh.close()


def _append_locked(results: list, balance_before: float, balance_after: float) -> Path:
    """
    Load, append, and atomically save PROD_LOG_PATH. Must only be called while
    holding the sidecar lock (see append_to_prod_log).

    Args:
        results (list): List of TradeResult objects from this run.
        balance_before (float): Account balance in dollars before this run's trades.
        balance_after (float): Account balance in dollars after this run's trades.

    Returns:
        Path: Absolute path to the trade log file (PROD_LOG_PATH).
    """
    if PROD_LOG_PATH.exists():
        wb = openpyxl.load_workbook(PROD_LOG_PATH)
        ws = wb.active
    else:
        # First run — create the workbook and apply the header row
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade Log"
        _apply_header_row(ws, _HEADER_FILL_PROD)

    # Use local time for the separator row so timestamps are human-readable
    run_ts = datetime.now(UTC).astimezone()
    _write_separator_row(ws, run_ts, balance_before, balance_after, len(results))
    _write_trade_rows(ws, results, run_ts)

    # Atomic save: openpyxl's wb.save() writes directly to the target path,
    # truncating it first — a crash partway through would destroy the entire
    # accumulated history. Writing to a tmp file and renaming it into place
    # means the visible PROD_LOG_PATH only ever transitions between complete
    # states (os.replace is an atomic rename on POSIX).
    tmp_path = PROD_LOG_PATH.with_name(PROD_LOG_PATH.name + ".tmp")
    wb.save(tmp_path)
    os.replace(tmp_path, PROD_LOG_PATH)
    logging.info("Trade log updated: %s (%d new row(s))", PROD_LOG_PATH, len(results))
    return PROD_LOG_PATH


def _write_fallback_log(results: list, balance_before: float, balance_after: float) -> Path:
    """
    Write this run's trade rows to a standalone timestamped file instead of the
    shared trade_log.xlsx.

    Called only when the sidecar lock could not be acquired within
    _LOCK_TIMEOUT_SECONDS — guarantees this run's rows are never silently lost,
    at the cost of splitting the trade history across an extra file the operator
    must merge in by hand (or simply keep alongside the main log).

    Args:
        results (list): List of TradeResult objects from this run.
        balance_before (float): Account balance in dollars before this run's trades.
        balance_after (float): Account balance in dollars after this run's trades.

    Returns:
        Path: Absolute path to the fallback file
            (PROJECT_ROOT / "trade_log_<timestamp>.xlsx").
    """
    run_ts = datetime.now(UTC).astimezone()
    fallback_path = PROJECT_ROOT / f"trade_log_{run_ts.strftime('%Y-%m-%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    _apply_header_row(ws, _HEADER_FILL_PROD)
    _write_separator_row(ws, run_ts, balance_before, balance_after, len(results))
    _write_trade_rows(ws, results, run_ts)

    wb.save(fallback_path)
    logging.info("Fallback trade log written: %s (%d row(s))", fallback_path, len(results))
    return fallback_path


# ─────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────

def append_to_prod_log(results: list, balance_before: float, balance_after: float) -> Path:
    """
    Append executed trade results to the persistent production trade log Excel file.

    If the file does not yet exist, creates it with a styled dark-blue header row.
    Each call appends a run-separator row (showing timestamp and balance change)
    followed by one data row per trade result, color-coded by status. The file is
    designed to accumulate all runs over the life of the bot.

    Concurrent callers (e.g. the scheduler and a manual run racing) are
    coordinated via a sidecar advisory lock (PROD_LOG_PATH + ".lock") so a
    load -> append -> save race can't silently clobber rows from another run,
    and the save itself is atomic (tmp file + os.replace) so a crash mid-save
    can't truncate the accumulated history. If the lock can't be acquired within
    _LOCK_TIMEOUT_SECONDS, this run's rows are never dropped — they're written
    to a separate timestamped fallback file instead (see _write_fallback_log).

    Args:
        results (list): List of TradeResult objects from trader.execute_trades().
            May be empty if no trades were executed this run.
        balance_before (float): Account balance in dollars before this run's trades.
        balance_after (float): Account balance in dollars after this run's trades.

    Returns:
        Path: Absolute path to the file actually written — either PROD_LOG_PATH
            on the normal path, or a standalone fallback file if the lock timed out.
    """
    lock_fh = _acquire_lock(_LOCK_PATH)
    if lock_fh is None:
        # Never clobber the shared log and never lose this run's rows: another
        # process appears to be mid-save. Write a standalone file instead.
        logging.warning(
            "Could not acquire lock on %s within %ds — writing this run's rows "
            "to a standalone fallback file instead of the shared trade log",
            _LOCK_PATH, _LOCK_TIMEOUT_SECONDS,
        )
        return _write_fallback_log(results, balance_before, balance_after)

    try:
        return _append_locked(results, balance_before, balance_after)
    finally:
        _release_lock(lock_fh)


def write_dev_simulation(
    results: list,
    all_candidates: list,
    balance_cents: int,
) -> Path:
    """
    Write a new timestamped dev simulation Excel file with two sheets.

    Sheet 1 ("Simulated Trades") contains trades that would have been executed
    if this were a real production run. Sheet 2 ("All Candidates") lists every
    qualifying pair discovered by the scanner — tradeable and non-tradeable alike —
    so the developer can see what the bot found before sizing decisions were applied.

    A new file is created on each dev run (never appended to) so simulation outputs
    are preserved for later comparison. The filename embeds the run timestamp.

    Args:
        results (list): List of TradeResult objects from trader.execute_trades()
            with status="simulated". May be empty.
        all_candidates (list): List of CandidatePair objects from scanner.py,
            representing all pairs discovered this run (regardless of tradeability).
        balance_cents (int): Virtual account balance in cents used for trade sizing.

    Returns:
        Path: Absolute path to the newly created simulation file
            (PROJECT_ROOT / "dev_simulation_YYYY-MM-DD_HHMMSS.xlsx").
    """
    run_ts   = datetime.now(UTC).astimezone()
    filename = f"dev_simulation_{run_ts.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    out_path = PROJECT_ROOT / filename

    wb = openpyxl.Workbook()

    # ── Sheet 1: Simulated Trades ──────────────────────────
    ws_trades = wb.active
    ws_trades.title = "Simulated Trades"
    _apply_header_row(ws_trades, _HEADER_FILL_DEV)

    # Run summary in row 2
    summary_cell = ws_trades.cell(
        row=2, column=1,
        value=(f"DEV SIMULATION  |  Run: {run_ts.strftime('%Y-%m-%d %H:%M')}  |  "
               f"Balance: ${balance_cents / 100:.2f}  |  "
               f"{len(results)} simulated trade(s)")
    )
    summary_cell.font = Font(italic=True, bold=True, color="375623", size=9)
    summary_cell.fill = PatternFill("solid", fgColor="E2EFDA")
    ws_trades.merge_cells(
        start_row=2, start_column=1,
        end_row=2, end_column=len(_TRADE_COLUMNS)
    )

    for result in results:
        row_idx = ws_trades.max_row + 1
        row_data = _result_to_row(result, run_ts)
        for col_idx, value in enumerate(row_data, start=1):
            ws_trades.cell(row=row_idx, column=col_idx, value=value)
        _apply_data_row_styles(ws_trades, row_idx, result.status)
        _apply_number_formats(ws_trades, row_idx)

    # ── Sheet 2: All Candidates ────────────────────────────
    ws_cands = wb.create_sheet("All Candidates")
    cand_headers = [
        ("Pair Type", 14),
        ("Market A", 45), ("Ticker A", 18),
        ("Market B", 45), ("Ticker B", 18),
        ("A Deadline", 12), ("B Deadline", 12),
        ("pA (YES ask)", 13), ("pB (YES ask)", 13), ("nA (NO ask)", 13),
        ("Price Diff", 12), ("Tradeable?", 12),
    ]
    for col_idx, (header, width) in enumerate(cand_headers, start=1):
        cell = ws_cands.cell(row=1, column=col_idx, value=header)
        cell.font  = Font(bold=True, color="FFFFFF", size=11)
        cell.fill  = PatternFill("solid", fgColor="375623")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_cands.column_dimensions[get_column_letter(col_idx)].width = width
    ws_cands.freeze_panes = "A2"
    ws_cands.row_dimensions[1].height = 28

    for pair in all_candidates:
        row_idx = ws_cands.max_row + 1
        diff    = pair.pA - pair.pB
        row_data = [
            pair.pair_type,
            display_title(pair.market_a), pair.market_a.ticker,
            display_title(pair.market_b), pair.market_b.ticker,
            pair.market_a.close_time.strftime("%Y-%m-%d") if pair.market_a.close_time else "",
            pair.market_b.close_time.strftime("%Y-%m-%d") if pair.market_b.close_time else "",
            round(pair.pA, 4),
            round(pair.pB, 4),
            round(pair.nA, 4),
            round(diff, 4),
            "YES" if pair.tradeable else "no",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_cands.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Highlight tradeable rows
        row_fill = PatternFill("solid", fgColor="E2EFDA" if pair.tradeable else "FFFFFF")
        for col in range(1, len(cand_headers) + 1):
            ws_cands.cell(row=row_idx, column=col).fill = row_fill

        # Format price columns (shifted right by 1 due to new Pair Type column)
        for col in (8, 9, 10, 11):
            ws_cands.cell(row=row_idx, column=col).number_format = "0.00%"

    wb.save(out_path)
    logging.info("Dev simulation written: %s", out_path)
    return out_path
